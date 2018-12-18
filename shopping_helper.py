from classes.item import item
from classes.shopping_list import shopping_list
from termcolor import colored, cprint
import openpyxl
import os.path
import tkinter as tk

def check_item(item_name, item_price):
	add = True
	for i in item_master:
		if i.name == item_name:
			i.add_price(item_price)
			return
	#if you got this far, item does not exist
	item_master.append(item(item_name,item_price))

def read_wb(sheet):
	end = 'B' + str(sheet.max_row)
	for row in sheet['A1': end]:
		name = row[0].value
		price = row[1].value
		check_item(name, price)

def output_items(item_list):
	for i in item_list:
		print("Name: " + i.name + " Avg: " + str(i.get_average()) + " Max: " + str(i.max_price)
				+ " Min: " + str(i.min_price))

def add_item_to_master():
	name = input("Please enter the name of the item you wish to add: ")
	price = eval(input("Please enter the price for " + name + ": "))
	check_item(name, price)

def add_to_shop_list(shop_list):
	name = input("Please enter the item name: ")
	for i in item_master:
		if i.name == name:
			shop_list.add_item(i)
			return
	cprint("Item not found", 'red', attrs=['bold'])

def remove_from_shop_list(shop_list):
	name = input("Please enter the item name: ")
	shop_list.remove_item(i)

def save_shop_list(shop_list, wb):
	sheet_list = wb.sheetnames

	try:
		idx = sheet_list.index(shop_list.name)
		sheet = wb[shop_list.name]
	except:
		cprint("Sheet does not exist, adding sheet " + shop_list.name, 'red', attrs=['bold'])
		sheet_name = shop_list.name
		sheet = wb.create_sheet(sheet_name)

	item_list = shop_list.item_list
	for i in range(1, len(item_list) + 1):
		sheet.cell(row=i, column=1, value=item_list[i - 1].name)

def load_shop_list(sheet):
	shop_list = shopping_list()
	shop_list.name = sheet.title
	end = 'B' + str(sheet.max_row)
	for row in sheet["A1":end]:
		item_name = row[0].value;
		for i in item_master:
			if i.name == item_name:
				shop_list.add_item(i)
				break

	return shop_list


def shopping_menu(shop_list):
	while True:
		cprint("\tNow editing " + shop_list.name + ":", attrs=['bold'])
		print("\t\t1. Add item to list")
		print("\t\t2. Remove item from list")
		print("\t\t3. Save list to data file")
		print("\t\t4. Output list")
		print("\t\t0. Go Back\n")

		try:
			choice = eval(input("Please enter your choice: "))
		except:
			cprint("That is not a correct choice", 'red', attrs=['bold'])
			continue
			cprint("\n\t--------------------\n", 'green' , attrs=['bold'])
		if choice == 1:
			add_to_shop_list(shop_list)
		elif choice == 2:
			remove_from_shop_list(shop_list)
		elif choice == 3:
			save_shop_list(shop_list)
		elif choice == 4:
			cprint("\n\tWhich do you want:", attrs=['bold'])
			print("\t\t1. Average prices")
			print("\t\t2. Maximum prices")
			print("\t\t3. Minimum prices\n")
			try:
				output_choice = eval(input("Please enter your choice: "))
				cprint("\n\t--------------------\n", 'green' , attrs=['bold'])
			except:
				cprint("That is not a correct choice", 'red', attrs=['bold'])
				continue

			if output_choice == 1:
				shop_list.output()
			elif output_choice == 2:
				shop_list.output_mode(0)
			elif output_choice == 3:
				shop_list.output_mode(1)
		elif choice == 0:
			return
		cprint("\n\t--------------------\n", 'green' , attrs=['bold'])

def list_lists():
	print("Now showing all shopping lists: ")
	print("\tTOTAL" + "\t NAME")
	for l in list_master:
		print("\t" + str(l.avg_total) + "\t" + l.name)

def save_file(wb):
	cprint("Saving file...", 'green')
	#saves all of the lists
	for sl in list_master:
		save_shop_list(sl, wb)

	wb.save(file_name)

item_master = []
list_master = []

file_name = input("\nPlease enter data file name: ")
if os.path.isfile(file_name): #if file exists, open it, else create it
	wb = openpyxl.load_workbook(file_name)

	#load all of the sheets in wb
	for sheet in wb:
		if sheet.title == 'Data':
			data_sheet = True
			read_wb(sheet)
		else:
			list_master.append(load_shop_list(sheet))
	if not data_sheet:
		cprint("'Data' sheet could not be found! Quitting...", 'red', attrs['bold'])
else:
	cprint("File not found, creating new file with name " + file_name, 'red')
	wb = openpyxl.Workbook()
	wb.create_sheet('Data', 0)

cprint("\n\t--------------------\n", 'green' , attrs=['bold'])
loop = True
while loop:
	cprint("\tWhat would you like to do?", attrs=['bold'])
	print("\t\t1. Add an item")
	print("\t\t2. Create a new shopping list")
	print("\t\t3. Edit an existing shopping list")
	print("\t\t4. Show all Shopping Lists")
	print("\t\t5. Output master item list")
	cprint("\t\t0. Quit\n", 'red')
	try:
		choice = eval(input("Please enter your choice: "))
	except:
		cprint("That is not a correct choice", 'red', attrs=['bold'])
		continue
	cprint("\n\t--------------------\n", 'green' , attrs=['bold'])
	if choice == 1:
		add_item_to_master()
	elif choice == 2:
		shop_list = shopping_list()
		while True:
			list_name = input("What would you like to name your list? ")
			good_name = True
			for l in list_master:
				if l.name == list_name:
					cprint("A list already exists with that name", 'red')
					good_name = False
			if good_name:
				shop_list.name = list_name
				break
		list_master.append(shop_list)
		shopping_menu(shop_list)
	elif choice == 3:
		shop_list_input = input("Please enter the name of the list: ")
		for l in list_master:
			if l.name == shop_list_input:
				shopping_menu(l)
				break
	elif choice == 4:
		list_lists()
	elif choice == 5:
		output_items(item_master)
	elif choice == 0:
		save_file(wb)
		cprint("\tThank you for using ShoppingHelper!", 'magenta', attrs=['bold'])
		loop = False
	cprint("\n\t--------------------\n", 'green' , attrs=['bold'])